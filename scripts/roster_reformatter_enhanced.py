import openpyxl
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime, time, timedelta
import re

def is_time(value):
    if isinstance(value, time):
        return True
    if isinstance(value, str):
        try:
            datetime.strptime(value, "%H:%M:%S")
            return True
        except ValueError:
            return False
    return False

def remove_non_numeric(s):
    return re.sub(r'[^0-9]', '', s)

def reformat_roster(file_path):
    print(f"Opening workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"Active worksheet: {ws.title}")

    # Step 1: Delete rows 1 through 5
    ws.delete_rows(1, 5)
    print("Step 1: Deleted rows 1 through 5")
    wb.save(file_path)

    # Step 2: Delete columns F through T
    ws.delete_cols(6, 15)
    print("Step 2: Deleted columns F through T")
    wb.save(file_path)

    # Step 3: Move column D to column A
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        value = ws.cell(row=row, column=4).value
        ws.cell(row=row, column=4).value = ws.cell(row=row, column=3).value
        ws.cell(row=row, column=3).value = ws.cell(row=row, column=2).value
        ws.cell(row=row, column=2).value = ws.cell(row=row, column=1).value
        ws.cell(row=row, column=1).value = value
    print("Step 3: Moved column D to column A")
    wb.save(file_path)

    # Step 4-8: Set header texts
    headers = ["SUBJECT", "START DATE", "START TIME", "END TIME", "DESCRIPTION"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    print("Step 4-8: Set header texts for columns A through E")
    wb.save(file_path)

    # Find the last row with data
    last_row = ws.max_row

    # Step 9: Fill blank cells in column A
    for row in range(2, last_row + 1):
        if ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value == "":
            ws.cell(row=row, column=1).value = ws.cell(row=row, column=3).value
    print("Step 9: Filled blank cells in column A with corresponding data from column C")
    wb.save(file_path)

    # Step 10: Handle STUD entries
    for row in range(2, last_row + 1):
        if ws.cell(row=row, column=1).value == "STUD":
            ws.cell(row=row, column=3).value = time(9, 0)
            ws.cell(row=row, column=4).value = time(17, 0)
    print("Step 10: Handled STUD entries")
    wb.save(file_path)

    # Step 11: Fill non-time cells in column C
    col_c = column_index_from_string('C')
    for row in range(2, last_row + 1):
        cell_value = ws.cell(row=row, column=col_c).value
        if not is_time(cell_value):
            ws.cell(row=row, column=col_c).value = time(0, 0)
            ws.cell(row=row, column=col_c).number_format = 'hh:mm'
    print("Step 11: Filled non-time cells in column C with 00:00")
    wb.save(file_path)

    # Step 12: Fill blank cells in column D with 23:59
    col_d = column_index_from_string('D')
    for row in range(2, last_row + 1):
        cell_value = ws.cell(row=row, column=col_d).value
        if cell_value is None or cell_value == "":
            ws.cell(row=row, column=col_d).value = time(23, 59)
            ws.cell(row=row, column=col_d).number_format = 'hh:mm'
    print("Step 12: Filled blank cells in column D with 23:59")
    wb.save(file_path)

    # Step 13: Fill blank cells in column E
    col_a = column_index_from_string('A')
    col_e = column_index_from_string('E')
    additional_text = "***RD = Rest Day, A/L = Annual Leave, Any days at the start or end of these days may also be impacted by shifts starting/ finishing during them, example may not finish work from previous work day until 02.00 but technically that day is still a RD or A/L***"
    for row in range(2, last_row + 1):
        cell_value_e = ws.cell(row=row, column=col_e).value
        if cell_value_e is None or cell_value_e == "":
            cell_value_a = ws.cell(row=row, column=col_a).value
            ws.cell(row=row, column=col_e).value = f"{cell_value_a} {additional_text}"
    print("Step 13: Filled blank cells in column E with data from column A and additional text")
    wb.save(file_path)

    # Step 14: Modify column B dates
    col_b = column_index_from_string('B')
    for row in range(2, last_row + 1):
        cell_value = ws.cell(row=row, column=col_b).value
        if cell_value:
            cell_value = str(cell_value)[6:]
            cell_value = remove_non_numeric(cell_value)
            if len(cell_value) == 8:
                try:
                    date_obj = datetime.strptime(cell_value, "%d%m%Y")
                    ws.cell(row=row, column=col_b).value = date_obj
                    ws.cell(row=row, column=col_b).number_format = 'DD/MM/YYYY'
                except ValueError:
                    print(f"Warning: Invalid date in row {row}")
    print("Step 14: Modified dates in column B")
    wb.save(file_path)

    # Step 15: Add END DATE column
    ws.insert_cols(4)
    ws.cell(row=1, column=4, value="END DATE")
    print("Step 15: Added new 'END DATE' column between C and D")
    ws.cell(row=1, column=5, value="END TIME")
    print("Step 15: Adjusted header for END TIME column")
    wb.save(file_path)

    # Step 16: Calculate END DATE
    col_b = column_index_from_string('B')
    col_c = column_index_from_string('C')
    col_d = column_index_from_string('D')
    col_e = column_index_from_string('E')
    for row in range(2, last_row + 1):
        start_date = ws.cell(row=row, column=col_b).value
        start_time = ws.cell(row=row, column=col_c).value
        end_time = ws.cell(row=row, column=col_e).value
        if isinstance(start_date, datetime) and isinstance(start_time, time) and isinstance(end_time, time):
            start_datetime = datetime.combine(start_date.date(), start_time)
            end_datetime = datetime.combine(start_date.date(), end_time)
            if end_time < start_time:
                end_datetime += timedelta(days=1)
            ws.cell(row=row, column=col_d).value = end_datetime.date()
            ws.cell(row=row, column=col_d).number_format = 'DD/MM/YYYY'
    print("Step 16: Calculated END DATE for all rows")
    wb.save(file_path)

    # Step 17: Adjust RD and A/L start times
    col_a = column_index_from_string('A')
    col_c = column_index_from_string('C')
    col_e = column_index_from_string('E')
    for row in range(3, last_row + 1):
        subject = ws.cell(row=row, column=col_a).value
        if subject in ["RD", "A/L"]:
            prev_end_time = ws.cell(row=row-1, column=col_e).value
            if isinstance(prev_end_time, time) and time(0, 0) < prev_end_time <= time(3, 15):
                new_start_time = (datetime.combine(datetime.min, prev_end_time) + timedelta(minutes=1)).time()
                ws.cell(row=row, column=col_c).value = new_start_time
                ws.cell(row=row, column=col_c).number_format = 'HH:MM'
            ws.cell(row=row, column=col_e).value = time(23, 59)
            ws.cell(row=row, column=col_e).number_format = 'HH:MM'
    print("Step 17: Adjusted start times for RD and A/L entries")
    wb.save(file_path)

    # Step 18: Format shift lengths in Column F
    for row in range(2, last_row + 1):
        cell = ws.cell(row=row, column=6)
        if is_time(cell.value):
            if isinstance(cell.value, str):
                time_obj = datetime.strptime(cell.value, "%H:%M:%S").time()
            else:
                time_obj = cell.value
            formatted_time = time_obj.strftime("%H:%M")
            cell.value = f"(hrs:min)shift length - {formatted_time}"
            cell.number_format = '@'
    print("Step 18: Formatted shift lengths in Column F")
    wb.save(file_path)

    print("All steps completed.")

if __name__ == "__main__":
    file_path = input("Enter the full path to your Excel file: ")
    reformat_roster(file_path)
    input("Press Enter to exit...")
